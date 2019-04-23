import requests, argparse, bs4, sys
from openpyxl import Workbook

global list_url
#list_url='https://vancouver.craigslist.org/search/sss?query=psw10'
#list_url='https://vancouver.craigslist.org/nvn/ele/d/surrey-south-71-surround-sound-home/6860891830.html'
items = ['']*10000
urls = ['']*10000
descriptions = ['']*10000

def write_to_excel(count):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Query"
    currentrow = 1

    print("Created file Craigslist.xlsx")
    print("Wrote %i rows" % count)
    print("Saving file")

    for i in range(0, count):
        #write item name to first column
        worksheet.cell(column = 1, row = currentrow).value = str(items[i])
        #write URL to second column
        worksheet.cell(column = 2, row = currentrow).value = str(urls[i])
        #write Description to third column
        worksheet.cell(column = 3, row = currentrow).value = str(descriptions[i])
        #increment excel row
        currentrow = currentrow+1

    workbook.save("Craigslist.xlsx")

def set_excel_titles(count):
    items[count] = "ITEM"
    urls[count] = "URL"
    descriptions[count] = "DESCRIPTION"
    count = count+1
    return count

def search_list(url,count):
    res = requests.get(url)
    res.raise_for_status()
    soup = bs4.BeautifulSoup(res.text, 'html.parser')

    #for each link in the search
    for a in soup.find_all("a", class_="result-title hdrlnk"):
        #print title
        items[count] = a.string
        #print(items[count])
        #print URL
        urls[count] = a.get('href')
        #print(urls[count])
        print_description(a.get('href'),count)
        count = count+1
    print("Found %i items" % count)
    return count

def strip(txt):
    #remove excess lines in text
    ret=""
    for l in txt.split("\n"):
        if l.strip()!='':
            ret += l + "\n"
    return ret

def print_description(url,count):
    res = requests.get(url)
    res.raise_for_status()
    soup = bs4.BeautifulSoup(res.text, 'html.parser')
    soup.find_all(attrs={"id": "postingbody"})
    #strip whitespace from the text
    desc = strip(soup.get_text())
    #remove the unwanted text
    #TODO: if no maps is provided fix this
    desc = desc[desc.find("google map")+12:desc.find("email to friend")]
    #remove the QR code link
    descriptions[count] = desc.replace('QR Code Link to This Post','')
    #print(descriptions[count])

def main(argv):
    search = ''
    list_url='https://vancouver.craigslist.org/search/sss?query='
    parser = argparse.ArgumentParser()
    parser.add_argument('-s', action="store", dest="query")
    search = str(parser.parse_args().query)
    #store query in list and remove spaces
    query = search.split()
    string = ''
    #store query into string and prepare URL
    for i  in range(0, len(query)):
        list_url+=query[i]
        if(i != (len(query)-1)):
                list_url+="+"

    count = 0
    count = set_excel_titles(count)
    count = search_list(list_url,count)
    write_to_excel(count)
    print("Done.")

if __name__=="__main__":
    main(sys.argv[1:])
